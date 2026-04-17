import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import json
import numpy as np
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler

# Page configuration
st.set_page_config(
    page_title="SIAMA Toolbox",
    page_icon="🎨",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        padding: 1rem;
        background: linear-gradient(90deg, #f0f8ff 0%, #e6f3ff 100%);
        border-radius: 10px;
        margin-bottom: 2rem;
    }
    .toolkit-card {
        padding: 1.5rem;
        border-radius: 10px;
        background-color: #f8f9fa;
        border-left: 5px solid #1f77b4;
        margin: 1rem 0;
    }
    .step-header {
        color: #2c5aa0;
        font-weight: bold;
        font-size: 1.2rem;
        margin-top: 1rem;
    }
    </style>
""", unsafe_allow_html=True)

# Initialize session state
if 'current_page' not in st.session_state:
    st.session_state.current_page = 'home'
if 'sit_data' not in st.session_state:
    st.session_state.sit_data = {'stakeholders': [], 'roles': {}}
if 'sat_data' not in st.session_state:
    st.session_state.sat_data = {'relationship_matrix': pd.DataFrame()}
if 'mat_data' not in st.session_state:
    st.session_state.mat_data = {}
if 'nature_of_craft' not in st.session_state:
    st.session_state.nature_of_craft = {'current_status': {}, 'desired_status': {}}
if 'projects' not in st.session_state:
    st.session_state.projects = {}
if 'current_project' not in st.session_state:
    st.session_state.current_project = None

# ----- Project persistence helpers -----
PROJECT_STATE_KEYS = ['sit_data', 'sat_data', 'mat_data', 'nature_of_craft']
LOCALSTORAGE_KEY = 'siama_projects_v1'


def _project_json_default(obj):
    if isinstance(obj, pd.DataFrame):
        return {"__dataframe__": True, "data": obj.to_dict(orient="records")}
    if isinstance(obj, pd.Series):
        return {"__series__": True, "data": obj.to_dict()}
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating,)):
        return float(obj)
    if isinstance(obj, datetime):
        return obj.isoformat()
    return str(obj)


def _project_json_hook(d):
    if isinstance(d, dict):
        if d.get("__dataframe__"):
            return pd.DataFrame(d.get("data", []))
        if d.get("__series__"):
            return pd.Series(d.get("data", {}))
    return d


def serialize_projects(projects_dict):
    return json.dumps(projects_dict, default=_project_json_default)


def deserialize_projects(s):
    try:
        return json.loads(s or "{}", object_hook=_project_json_hook)
    except Exception:
        return {}


def snapshot_current_state():
    return {k: st.session_state.get(k) for k in PROJECT_STATE_KEYS}


def apply_state(state):
    for k in PROJECT_STATE_KEYS:
        if k in state:
            st.session_state[k] = state[k]


def reset_state():
    st.session_state.sit_data = {'stakeholders': [], 'roles': {}}
    st.session_state.sat_data = {'relationship_matrix': pd.DataFrame()}
    st.session_state.mat_data = {}
    st.session_state.nature_of_craft = {'current_status': {}, 'desired_status': {}}


# One-time hydration: if the URL carries a ?__siama_load=... payload from a
# "Load from browser" bridge, decode it into session state.
try:
    _qp = st.query_params
    _qp_get = lambda k: _qp.get(k)
    _qp_del = lambda k: _qp.pop(k, None) if hasattr(_qp, "pop") else _qp.__delitem__(k)
except AttributeError:
    _qp_get = lambda k: (st.experimental_get_query_params().get(k) or [None])[0]
    _qp_del = lambda k: st.experimental_set_query_params()

_hydrate_payload = _qp_get("__siama_load")
if _hydrate_payload:
    try:
        import base64 as _b64
        raw = _hydrate_payload if isinstance(_hydrate_payload, str) else _hydrate_payload[0]
        padded = raw + "=" * ((4 - len(raw) % 4) % 4)
        decoded = _b64.urlsafe_b64decode(padded.encode("ascii")).decode("utf-8")
        loaded = deserialize_projects(decoded)
        if isinstance(loaded, dict):
            st.session_state.projects = loaded
            st.session_state._loaded_from_browser = len(loaded)
    except Exception as _e:
        st.session_state._load_error = str(_e)
    try:
        _qp_del("__siama_load")
    except Exception:
        pass


def push_to_browser():
    import streamlit.components.v1 as components
    payload = serialize_projects(st.session_state.projects)
    js_literal = json.dumps(payload)
    components.html(
        """
        <div style="font: 13px -apple-system, BlinkMacSystemFont, sans-serif; color:#2c5aa0;">
          <span id="siama-save-status">⏳ Syncing to browser…</span>
        </div>
        <script>
          (function() {
            try {
              var target = window.top || window.parent;
              target.localStorage.setItem("__KEY__", __PAYLOAD__);
              document.getElementById("siama-save-status").innerText =
                "✅ Saved to browser at " + new Date().toLocaleTimeString();
            } catch (e) {
              document.getElementById("siama-save-status").innerText =
                "❌ Browser save failed: " + e.message;
            }
          })();
        </script>
        """.replace("__KEY__", LOCALSTORAGE_KEY).replace("__PAYLOAD__", js_literal),
        height=40,
    )


def pull_from_browser():
    import streamlit.components.v1 as components
    components.html(
        """
        <div style="font: 13px -apple-system, BlinkMacSystemFont, sans-serif; color:#2c5aa0;">
          <span id="siama-load-status">⏳ Reading browser storage…</span>
        </div>
        <script>
          (function() {
            try {
              var target = window.top || window.parent;
              var data = target.localStorage.getItem("__KEY__");
              if (!data) {
                document.getElementById("siama-load-status").innerText =
                  "ℹ️ No saved projects found in this browser.";
                return;
              }
              var b64 = btoa(unescape(encodeURIComponent(data)))
                .replace(/\\+/g,'-').replace(/\\//g,'_').replace(/=+$/,'');
              var url = new URL(target.location.href);
              url.searchParams.set("__siama_load", b64);
              target.location.replace(url.toString());
            } catch (e) {
              document.getElementById("siama-load-status").innerText =
                "❌ Browser load failed: " + e.message;
            }
          })();
        </script>
        """.replace("__KEY__", LOCALSTORAGE_KEY),
        height=40,
    )

# Navigation
st.sidebar.title("🎨 SIAMA Toolbox")
if st.session_state.get('current_project'):
    st.sidebar.caption(f"📁 **{st.session_state.current_project}**")
else:
    st.sidebar.caption("📁 *No project loaded*")
st.sidebar.markdown("---")

menu = st.sidebar.radio(
    "Navigate",
    ["🏠 Home", "📁 Projects", "1️⃣ SIT - Stakeholder Identification", "2️⃣ SAT - Stakeholder Analysis",
     "3️⃣ MAT - Market Analysis", "4️⃣ Nature of Craft", "📊 Summary & Export"]
)

# Home Page
if menu == "🏠 Home":
    st.markdown('<div class="main-header">SIAMA Toolbox</div>', unsafe_allow_html=True)
    st.markdown("### Stakeholder Identification, Analysis, and Market Assessment")
    
    st.markdown("""
    <div class="toolkit-card">
    <h3>🎯 Purpose</h3>
    The SIAMA toolbox is a comprehensive framework designed to help craft trainers and training organizations 
    systematically plan craft education programs by understanding contextually different needs of artisans in India.
    </div>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("""
        <div class="toolkit-card">
        <h3>❓ What should we teach?</h3>
        Addressed through:
        <ul>
        <li>Market study components (MAT)</li>
        <li>Artisan's prior skill, knowledge, and aspirations (SIT, SAT)</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="toolkit-card">
        <h3>👥 Whom should we teach?</h3>
        Addressed through:
        <ul>
        <li>Stakeholder identification (SIT)</li>
        <li>Stakeholder analysis (SAT)</li>
        </ul>
        </div>
        """, unsafe_allow_html=True)
    
    st.markdown("---")
    st.subheader("📦 The Three Toolkits")

    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("""
        **1. Stakeholder Identification Toolkit (SIT)**
        - 4 Sequential Steps
        - Identifies craft value chain actors
        - Maps stakeholder relationships
        - Visualizes supply chain
        """)

    with col2:
        st.markdown("""
        **2. Stakeholder Analysis Toolkit (SAT)**
        - 5 Sequential Tools
        - Analyzes power, interest, legitimacy
        - Conflict resolution strategies
        - Knowledge & responsibility mapping
        """)

    with col3:
        st.markdown("""
        **3. Market Analysis Toolkit (MAT)**
        - 8 Analysis Tools
        - Industry understanding
        - Customer insights
        - Brand assessment
        """)

    st.info("👈 Use the sidebar to navigate through each toolkit")

# SIT - Stakeholder Identification Toolkit
elif menu == "1️⃣ SIT - Stakeholder Identification":
    st.markdown('<div class="main-header">SIT - Stakeholder Identification Toolkit</div>', unsafe_allow_html=True)
    
    st.markdown("""
    SIT helps understand a craft's value chain by recognizing all individuals, groups, and organizations 
    that have a direct or indirect impact on it.
    """)
    
    tab1, tab2, tab3, tab4 = st.tabs([
        "Step 1: Predefined Roles & Questionnaires",
        "Step 2: Role × Actor Database",
        "Step 3: Role Card",
        "Step 4: Role Map"
    ])
    
    with tab1:
        st.markdown('<p class="step-header">Step 1: Predefined Roles and Questionnaires</p>', unsafe_allow_html=True)
        st.info("This tool provides predefined roles relevant to craft supply chains with structured questionnaires.")
        
        roles = ["Supplier", "Producer", "Refiner", "Marketer", "Buyer"]
        
        selected_role = st.selectbox("Select Role to Interview", roles)
        
        st.subheader(f"Primary Questionnaire for {selected_role}")
        
        questionnaires = {
            "Supplier": [
                "Who provides the raw materials?",
                "Where do they source materials from?",
                "What is their relationship with producers?"
            ],
            "Producer": [
                "Who creates the craft products?",
                "What skills do they possess?",
                "How long have they been practicing?"
            ],
            "Refiner": [
                "Who adds value to the basic product?",
                "What refinement processes are used?",
                "What expertise do they bring?"
            ],
            "Marketer": [
                "Who promotes the products?",
                "What channels do they use?",
                "What is their reach?"
            ],
            "Buyer": [
                "Who are the end consumers?",
                "What are their preferences?",
                "What price points do they prefer?"
            ]
        }
        
        responses = {}
        for q in questionnaires[selected_role]:
            responses[q] = st.text_area(q, key=f"q_{selected_role}_{q}")
        
        st.subheader("Secondary Questionnaire (Deep Dive)")
        secondary_questions = [
            "How frequently do they interact?",
            "What are the payment terms?",
            "Are there any challenges in this relationship?"
        ]
        
        for sq in secondary_questions:
            responses[sq] = st.text_area(sq, key=f"sq_{selected_role}_{sq}")
        
        if st.button("Save Stakeholder Data"):
            stakeholder_entry = {
                "role": selected_role,
                "responses": responses,
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            st.session_state.sit_data['stakeholders'].append(stakeholder_entry)
            st.success(f"✅ Data saved for {selected_role}")
    
    with tab2:
        st.markdown('<p class="step-header">Step 2: Role × Actor Database</p>', unsafe_allow_html=True)
        st.info("Record structured information about identified stakeholders.")
        
        if len(st.session_state.sit_data['stakeholders']) > 0:
            # Create a dataframe from stakeholders
            data_for_df = []
            for s in st.session_state.sit_data['stakeholders']:
                row = {"Role": s['role'], "Timestamp": s['timestamp']}
                for q, a in s['responses'].items():
                    row[q[:30]] = a[:50] if a else ""
                data_for_df.append(row)

            df = pd.DataFrame(data_for_df)
            st.dataframe(df, use_container_width=True)

            with st.expander("🗑️ Manage stakeholder entries"):
                for idx, s in enumerate(list(st.session_state.sit_data['stakeholders'])):
                    c1, c2 = st.columns([5, 1])
                    with c1:
                        st.write(f"**{idx+1}.** {s['role']} — {s['timestamp']}")
                    with c2:
                        if st.button("🗑️ Delete", key=f"del_sit_sh_{idx}"):
                            st.session_state.sit_data['stakeholders'].pop(idx)
                            st.rerun()
            
            # Edit/Add actors
            st.subheader("Add Actor Details")
            col1, col2 = st.columns(2)
            with col1:
                actor_name = st.text_input("Actor Name")
                actor_role = st.selectbox("Role", ["Supplier", "Producer", "Refiner", "Marketer", "Buyer"])
            with col2:
                actor_location = st.text_input("Location")
                actor_contact = st.text_input("Contact Information")
            
            actor_details = st.text_area("Additional Details")
            
            if st.button("Add Actor"):
                if actor_name:
                    if actor_role not in st.session_state.sit_data['roles']:
                        st.session_state.sit_data['roles'][actor_role] = []
                    
                    st.session_state.sit_data['roles'][actor_role].append({
                        "name": actor_name,
                        "location": actor_location,
                        "contact": actor_contact,
                        "details": actor_details
                    })
                    st.success(f"✅ Actor {actor_name} added to {actor_role}")
        else:
            st.warning("⚠️ Please complete Step 1 first to collect stakeholder data.")
    
    with tab3:
        st.markdown('<p class="step-header">Step 3: Role Card Visualization</p>', unsafe_allow_html=True)
        st.info("Visualize stakeholders clustered around specific roles in the supply chain.")
        
        if st.session_state.sit_data['roles']:
            for role in list(st.session_state.sit_data['roles'].keys()):
                actors = st.session_state.sit_data['roles'][role]
                with st.expander(f"📋 {role} ({len(actors)} actors)"):
                    for i, actor in enumerate(list(actors), 1):
                        c1, c2 = st.columns([6, 1])
                        with c1:
                            st.markdown(f"""
                            **{i}. {actor['name']}**
                            - Location: {actor['location']}
                            - Contact: {actor['contact']}
                            - Details: {actor['details']}
                            """)
                        with c2:
                            if st.button("🗑️", key=f"del_actor_{role}_{i-1}", help=f"Delete {actor['name']}"):
                                st.session_state.sit_data['roles'][role].pop(i-1)
                                st.rerun()
                    st.markdown("---")
                    rc1, rc2 = st.columns([3, 1])
                    with rc2:
                        if st.button(f"🗑️ Delete role '{role}'", key=f"del_role_{role}"):
                            del st.session_state.sit_data['roles'][role]
                            st.rerun()
            
            # Simple visualization
            role_counts = {role: len(actors) for role, actors in st.session_state.sit_data['roles'].items()}
            fig = px.bar(
                x=list(role_counts.keys()),
                y=list(role_counts.values()),
                labels={'x': 'Role', 'y': 'Number of Actors'},
                title='Stakeholder Distribution by Role'
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.warning("⚠️ Please add actors in Step 2 first.")
    
    with tab4:
        st.markdown('<p class="step-header">Step 4: Role Map</p>', unsafe_allow_html=True)
        st.info("Visualize the entire supply chain with flow markers showing how resources move.")
        
        if st.session_state.sit_data['roles']:
            st.subheader("Supply Chain Flow")
            
            # Create a network diagram
            fig = go.Figure()
            
            roles = ["Supplier", "Producer", "Refiner", "Marketer", "Buyer"]
            x_pos = [i for i in range(len(roles))]
            
            # Add nodes
            for i, role in enumerate(roles):
                actor_count = len(st.session_state.sit_data['roles'].get(role, []))
                fig.add_trace(go.Scatter(
                    x=[i], y=[0],
                    mode='markers+text',
                    marker=dict(size=40 + actor_count*10, color='lightblue'),
                    text=f"{role}<br>({actor_count})",
                    textposition="top center",
                    name=role
                ))
            
            # Add arrows
            for i in range(len(roles)-1):
                fig.add_annotation(
                    x=i+0.5, y=0,
                    ax=i, ay=0,
                    xref='x', yref='y',
                    axref='x', ayref='y',
                    text='',
                    showarrow=True,
                    arrowhead=2,
                    arrowsize=1.5,
                    arrowwidth=2,
                    arrowcolor='gray'
                )
            
            fig.update_layout(
                title='Craft Value Chain Flow',
                showlegend=False,
                height=400,
                xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
                yaxis=dict(showgrid=False, zeroline=False, showticklabels=False)
            )
            
            st.plotly_chart(fig, use_container_width=True)
            
            st.markdown("### Flow Indicators")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Product Flow", "→", delta="Supplier to Buyer")
            with col2:
                st.metric("Money Flow", "←", delta="Buyer to Supplier")
            with col3:
                st.metric("Information Flow", "↔", delta="Bidirectional")
        else:
            st.warning("⚠️ Please complete previous steps to visualize the role map.")

# SAT - Stakeholder Analysis Toolkit
elif menu == "2️⃣ SAT - Stakeholder Analysis":
    st.markdown('<div class="main-header">SAT - Stakeholder Analysis Toolkit</div>', unsafe_allow_html=True)
    
    st.markdown("""
    SAT evaluates the roles, interests, and influences of identified stakeholders to tailor training programs effectively.
    """)
    
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "Step 1: Relationship Matrix",
        "Step 2: Management Tool",
        "Step 3: Conflict Resolution",
        "Step 4: Knowledge & Responsibility",
        "Step 5: Value Exchange Map"
    ])
    
    with tab1:
        st.markdown('<p class="step-header">Step 1: Stakeholder Relationship Matrix</p>', unsafe_allow_html=True)
        st.info("Rate each stakeholder on Power, Interest, Legitimacy, and Urgency.")
        
        if st.session_state.sit_data['roles']:
            # Get all stakeholders
            all_stakeholders = []
            for role, actors in st.session_state.sit_data['roles'].items():
                for actor in actors:
                    all_stakeholders.append(f"{actor['name']} ({role})")
            
            if all_stakeholders:
                st.subheader("Rate Stakeholders")
                
                selected_stakeholder = st.selectbox("Select Stakeholder to Rate", all_stakeholders)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    power = st.slider("Power (ability to influence outcomes)", 1, 10, 5)
                    interest = st.slider("Interest (level of concern/stake)", 1, 10, 5)
                
                with col2:
                    legitimacy = st.slider("Legitimacy (validity of involvement)", 1, 10, 5)
                    urgency = st.slider("Urgency (immediacy of demands)", 1, 10, 5)
                
                st.subheader("Relationship Details")
                interactions = st.text_area("How do they interact with others?")
                tasks = st.text_area("What tasks do they perform?")
                knowledge = st.text_area("What knowledge/skills do they share?")
                
                if st.button("Save Rating"):
                    if 'relationship_data' not in st.session_state.sat_data:
                        st.session_state.sat_data['relationship_data'] = []
                    
                    st.session_state.sat_data['relationship_data'].append({
                        'stakeholder': selected_stakeholder,
                        'power': power,
                        'interest': interest,
                        'legitimacy': legitimacy,
                        'urgency': urgency,
                        'interactions': interactions,
                        'tasks': tasks,
                        'knowledge': knowledge
                    })
                    st.success("✅ Rating saved!")
                
                # Display existing ratings
                if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
                    st.subheader("Current Ratings")
                    df = pd.DataFrame(st.session_state.sat_data['relationship_data'])
                    st.dataframe(df, use_container_width=True)

                    with st.expander("🗑️ Manage rating entries"):
                        for idx, r in enumerate(list(st.session_state.sat_data['relationship_data'])):
                            c1, c2 = st.columns([5, 1])
                            with c1:
                                st.write(f"**{idx+1}.** {r.get('stakeholder','?')} — "
                                         f"P:{r.get('power','-')}, I:{r.get('interest','-')}, "
                                         f"L:{r.get('legitimacy','-')}, U:{r.get('urgency','-')}")
                            with c2:
                                if st.button("🗑️ Delete", key=f"del_sat_rat_{idx}"):
                                    removed = st.session_state.sat_data['relationship_data'].pop(idx)
                                    # Also clean up subgroup memberships referencing this stakeholder
                                    sh_name = removed.get('stakeholder')
                                    for sg in st.session_state.sat_data.get('subgroups', {}).values():
                                        if sh_name in sg.get('members', []):
                                            sg['members'].remove(sh_name)
                                    st.session_state.sat_data.get('subgroup_assignments', {}).pop(sh_name, None)
                                    st.rerun()
        else:
            st.warning("⚠️ Please complete SIT first to identify stakeholders.")
    
    with tab2:
        st.markdown('<p class="step-header">Step 2: Stakeholder Management Tool</p>', unsafe_allow_html=True)
        st.info("Visualize stakeholders on Power vs Interest, Power vs Legitimacy, and Power vs Urgency. Create subgroups for detailed analysis.")

        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            # Add tabs for main analysis and subgroup creation
            subtab1, subtab2 = st.tabs(["📊 Stakeholder Analysis", "👥 Subgroup Management"])

            with subtab1:
                df = pd.DataFrame(st.session_state.sat_data['relationship_data'])

                chart_type = st.selectbox("Select Comparison",
                                         ["Power vs Interest", "Power vs Legitimacy", "Power vs Urgency"])

                if chart_type == "Power vs Interest":
                    fig = px.scatter(df, x='power', y='interest',
                                   text='stakeholder',
                                   title='Power vs Interest Matrix')
                elif chart_type == "Power vs Legitimacy":
                    fig = px.scatter(df, x='power', y='legitimacy',
                                   text='stakeholder',
                                   title='Power vs Legitimacy Matrix')
                else:
                    fig = px.scatter(df, x='power', y='urgency',
                                   text='stakeholder',
                                   title='Power vs Urgency Matrix')

                fig.update_traces(textposition='top center')
                fig.update_layout(height=500)

                # Add quadrant lines
                fig.add_hline(y=5, line_dash="dash", line_color="gray")
                fig.add_vline(x=5, line_dash="dash", line_color="gray")

                st.plotly_chart(fig, use_container_width=True)

                # Strategy recommendations
                st.subheader("Management Strategies")
                col1, col2 = st.columns(2)

                with col1:
                    st.markdown("""
                    **High Power, High Interest:**
                    - Manage Closely
                    - Key Players
                    """)
                    st.markdown("""
                    **High Power, Low Interest:**
                    - Keep Satisfied
                    - Important but passive
                    """)

                with col2:
                    st.markdown("""
                    **Low Power, High Interest:**
                    - Keep Informed
                    - Show consideration
                    """)
                    st.markdown("""
                    **Low Power, Low Interest:**
                    - Monitor
                    - Minimal effort
                    """)

            with subtab2:
                st.subheader("Create and Manage Stakeholder Subgroups")
                st.info("Organize stakeholders into subgroups for targeted analysis and strategy development.")

                # Initialize subgroup data if not exists
                if 'subgroups' not in st.session_state.sat_data:
                    st.session_state.sat_data['subgroups'] = {}
                if 'subgroup_assignments' not in st.session_state.sat_data:
                    st.session_state.sat_data['subgroup_assignments'] = {}

                # Add clustering option
                clustering_tabs = st.tabs(["📝 Manual Grouping", "🤖 Automatic Clustering"])

                with clustering_tabs[0]:
                    st.markdown("### Manual Subgroup Creation")

                    # Section 1: Create/Manage Subgroups
                    col1, col2 = st.columns([2, 1])

                    with col1:
                        st.markdown("#### Define Subgroups")
                        subgroup_name = st.text_input("Subgroup Name", placeholder="e.g., Local Artisans, Urban Buyers, etc.")
                        subgroup_description = st.text_area("Description", placeholder="Describe the purpose of this subgroup")

                    with col2:
                        st.markdown("#### Actions")
                        if st.button("➕ Add Subgroup", use_container_width=True):
                            if subgroup_name:
                                if subgroup_name not in st.session_state.sat_data['subgroups']:
                                    st.session_state.sat_data['subgroups'][subgroup_name] = {
                                        'description': subgroup_description,
                                        'members': []
                                    }
                                    st.success(f"✅ Subgroup '{subgroup_name}' created!")
                                else:
                                    st.warning("⚠️ Subgroup already exists!")
                            else:
                                st.error("Please enter a subgroup name")

                    # Display existing subgroups
                    if st.session_state.sat_data['subgroups']:
                        st.markdown("---")
                        st.markdown("#### Existing Subgroups")

                        subgroups_df = pd.DataFrame([
                            {
                                'Subgroup': name,
                                'Description': data['description'],
                                'Members': len(data['members'])
                            }
                            for name, data in st.session_state.sat_data['subgroups'].items()
                        ])
                        st.dataframe(subgroups_df, use_container_width=True)

                        with st.expander("🗑️ Delete a subgroup"):
                            sg_names = list(st.session_state.sat_data['subgroups'].keys())
                            sg_to_delete = st.selectbox("Select subgroup to delete", sg_names, key="_sg_del_sel")
                            if st.button("🗑️ Delete selected subgroup", key="_sg_del_btn"):
                                if sg_to_delete and sg_to_delete in st.session_state.sat_data['subgroups']:
                                    # Clear member assignments pointing at this subgroup
                                    assignments = st.session_state.sat_data.get('subgroup_assignments', {})
                                    for member, sg in list(assignments.items()):
                                        if sg == sg_to_delete:
                                            del assignments[member]
                                    del st.session_state.sat_data['subgroups'][sg_to_delete]
                                    st.success(f"Deleted subgroup '{sg_to_delete}'.")
                                    st.rerun()

                    # Section 2: Assign Stakeholders to Subgroups
                    st.markdown("---")
                    st.markdown("#### Assign Stakeholders to Subgroups")

                    if st.session_state.sat_data['subgroups']:
                        col1, col2 = st.columns(2)

                        with col1:
                            stakeholders = [item['stakeholder'] for item in st.session_state.sat_data['relationship_data']]
                            selected_stakeholder = st.selectbox("Select Stakeholder", stakeholders)

                        with col2:
                            subgroup_options = list(st.session_state.sat_data['subgroups'].keys())
                            selected_subgroup = st.selectbox("Assign to Subgroup", subgroup_options)

                        if st.button("Assign Stakeholder to Subgroup"):
                            # Add stakeholder to subgroup
                            if selected_stakeholder not in st.session_state.sat_data['subgroups'][selected_subgroup]['members']:
                                st.session_state.sat_data['subgroups'][selected_subgroup]['members'].append(selected_stakeholder)
                                st.session_state.sat_data['subgroup_assignments'][selected_stakeholder] = selected_subgroup
                                st.success(f"✅ {selected_stakeholder} assigned to {selected_subgroup}")
                            else:
                                st.info(f"ℹ️ {selected_stakeholder} is already in {selected_subgroup}")

                        # Display current assignments
                        st.markdown("---")
                        st.markdown("#### Current Assignments")

                        for subgroup_name, subgroup_data in st.session_state.sat_data['subgroups'].items():
                            if subgroup_data['members']:
                                with st.expander(f"📁 {subgroup_name} ({len(subgroup_data['members'])} members)"):
                                    for member in subgroup_data['members']:
                                        col1, col2 = st.columns([3, 1])
                                        with col1:
                                            st.write(f"• {member}")
                                        with col2:
                                            if st.button(f"Remove", key=f"remove_{subgroup_name}_{member}"):
                                                st.session_state.sat_data['subgroups'][subgroup_name]['members'].remove(member)
                                                if member in st.session_state.sat_data['subgroup_assignments']:
                                                    del st.session_state.sat_data['subgroup_assignments'][member]
                                                st.rerun()

                with clustering_tabs[1]:
                    st.markdown("### Automatic K-Means Clustering")
                    st.info("Use machine learning to automatically group stakeholders based on their Power, Interest, Legitimacy, and Urgency ratings.")

                    # Get stakeholder data
                    df = pd.DataFrame(st.session_state.sat_data['relationship_data'])

                    col1, col2 = st.columns(2)

                    with col1:
                        num_clusters = st.slider("Number of Clusters", min_value=2, max_value=min(8, len(df)), value=3)
                        st.caption(f"Split {len(df)} stakeholders into {num_clusters} groups")

                    with col2:
                        features_to_use = st.multiselect(
                            "Features for Clustering",
                            ["power", "interest", "legitimacy", "urgency"],
                            default=["power", "interest", "legitimacy", "urgency"]
                        )

                    cluster_naming = st.selectbox(
                        "Cluster Naming Strategy",
                        ["Automatic (based on characteristics)", "Custom (manual naming)"]
                    )

                    if st.button("🤖 Generate Clusters", type="primary", use_container_width=True):
                        if len(features_to_use) < 2:
                            st.error("Please select at least 2 features for clustering")
                        else:
                            # Prepare data for clustering
                            X = df[features_to_use].values

                            # Standardize features
                            scaler = StandardScaler()
                            X_scaled = scaler.fit_transform(X)

                            # Perform K-means clustering
                            kmeans = KMeans(n_clusters=num_clusters, random_state=42, n_init=10)
                            df['cluster'] = kmeans.fit_predict(X_scaled)

                            # Calculate cluster centers in original scale
                            cluster_centers = scaler.inverse_transform(kmeans.cluster_centers_)

                            # Generate cluster names based on characteristics
                            cluster_names = {}
                            for i in range(num_clusters):
                                cluster_df = df[df['cluster'] == i]
                                avg_power = cluster_df['power'].mean()
                                avg_interest = cluster_df['interest'].mean()

                                # Name based on power-interest matrix
                                if avg_power > 6.5 and avg_interest > 6.5:
                                    name = f"Key Players (Cluster {i+1})"
                                    desc = "High power, high interest - manage closely"
                                elif avg_power > 6.5 and avg_interest <= 6.5:
                                    name = f"Keep Satisfied (Cluster {i+1})"
                                    desc = "High power, lower interest - keep satisfied"
                                elif avg_power <= 6.5 and avg_interest > 6.5:
                                    name = f"Keep Informed (Cluster {i+1})"
                                    desc = "Lower power, high interest - keep informed"
                                else:
                                    name = f"Monitor (Cluster {i+1})"
                                    desc = "Lower power, lower interest - monitor"

                                cluster_names[i] = {'name': name, 'description': desc}

                            # Create/update subgroups based on clusters
                            for i in range(num_clusters):
                                cluster_df = df[df['cluster'] == i]
                                members = cluster_df['stakeholder'].tolist()

                                subgroup_name = cluster_names[i]['name']
                                subgroup_desc = cluster_names[i]['description']

                                # Add cluster info to description
                                full_desc = f"{subgroup_desc}\n\nAuto-generated using K-means clustering with {', '.join(features_to_use)}"

                                st.session_state.sat_data['subgroups'][subgroup_name] = {
                                    'description': full_desc,
                                    'members': members
                                }

                                # Update assignments
                                for member in members:
                                    st.session_state.sat_data['subgroup_assignments'][member] = subgroup_name

                            st.success(f"✅ Successfully created {num_clusters} subgroups using K-means clustering!")
                            st.rerun()

                    # Display clustering visualization if clusters exist
                    if 'subgroups' in st.session_state.sat_data and st.session_state.sat_data['subgroups']:
                        st.markdown("---")
                        st.markdown("#### Clustering Visualization")

                        # Check if we have cluster assignments
                        df_viz = pd.DataFrame(st.session_state.sat_data['relationship_data'])
                        df_viz['Subgroup'] = df_viz['stakeholder'].map(st.session_state.sat_data.get('subgroup_assignments', {}))

                        if not df_viz['Subgroup'].isna().all():
                            # 3D scatter plot
                            st.markdown("##### 3D Cluster Visualization")

                            fig = px.scatter_3d(
                                df_viz,
                                x='power',
                                y='interest',
                                z='legitimacy',
                                color='Subgroup',
                                hover_data=['stakeholder', 'urgency'],
                                title='Stakeholder Clusters in 3D Space',
                                labels={'power': 'Power', 'interest': 'Interest', 'legitimacy': 'Legitimacy'}
                            )
                            fig.update_layout(height=600)
                            st.plotly_chart(fig, use_container_width=True)

                            # 2D projections
                            st.markdown("##### 2D Cluster Projections")

                            col1, col2 = st.columns(2)

                            with col1:
                                fig1 = px.scatter(
                                    df_viz,
                                    x='power',
                                    y='interest',
                                    color='Subgroup',
                                    hover_data=['stakeholder'],
                                    title='Power vs Interest'
                                )
                                fig1.add_hline(y=5, line_dash="dash", line_color="gray", opacity=0.5)
                                fig1.add_vline(x=5, line_dash="dash", line_color="gray", opacity=0.5)
                                st.plotly_chart(fig1, use_container_width=True)

                            with col2:
                                fig2 = px.scatter(
                                    df_viz,
                                    x='legitimacy',
                                    y='urgency',
                                    color='Subgroup',
                                    hover_data=['stakeholder'],
                                    title='Legitimacy vs Urgency'
                                )
                                fig2.add_hline(y=5, line_dash="dash", line_color="gray", opacity=0.5)
                                fig2.add_vline(x=5, line_dash="dash", line_color="gray", opacity=0.5)
                                st.plotly_chart(fig2, use_container_width=True)

                            # Cluster statistics
                            st.markdown("##### Cluster Statistics")

                            cluster_stats = []
                            for subgroup_name, subgroup_data in st.session_state.sat_data['subgroups'].items():
                                if subgroup_data['members']:
                                    members_df = df_viz[df_viz['stakeholder'].isin(subgroup_data['members'])]
                                    cluster_stats.append({
                                        'Subgroup': subgroup_name,
                                        'Count': len(subgroup_data['members']),
                                        'Avg Power': f"{members_df['power'].mean():.2f}",
                                        'Avg Interest': f"{members_df['interest'].mean():.2f}",
                                        'Avg Legitimacy': f"{members_df['legitimacy'].mean():.2f}",
                                        'Avg Urgency': f"{members_df['urgency'].mean():.2f}"
                                    })

                            if cluster_stats:
                                stats_df = pd.DataFrame(cluster_stats)
                                st.dataframe(stats_df, use_container_width=True)

                # Section 3: Subgroup Analysis (common for both manual and automatic)
                st.markdown("---")
                st.markdown("### Subgroup Analysis")

                if st.session_state.sat_data['subgroups']:
                    if any(data['members'] for data in st.session_state.sat_data['subgroups'].values()):
                        analysis_subgroup = st.selectbox(
                            "Select Subgroup to Analyze",
                            [name for name, data in st.session_state.sat_data['subgroups'].items() if data['members']]
                        )

                        if analysis_subgroup:
                            members = st.session_state.sat_data['subgroups'][analysis_subgroup]['members']

                            # Get ratings for subgroup members
                            df = pd.DataFrame(st.session_state.sat_data['relationship_data'])
                            subgroup_df = df[df['stakeholder'].isin(members)]

                            if not subgroup_df.empty:
                                # Calculate aggregated metrics
                                col1, col2, col3, col4 = st.columns(4)

                                with col1:
                                    avg_power = subgroup_df['power'].mean()
                                    st.metric("Avg Power", f"{avg_power:.1f}")

                                with col2:
                                    avg_interest = subgroup_df['interest'].mean()
                                    st.metric("Avg Interest", f"{avg_interest:.1f}")

                                with col3:
                                    avg_legitimacy = subgroup_df['legitimacy'].mean()
                                    st.metric("Avg Legitimacy", f"{avg_legitimacy:.1f}")

                                with col4:
                                    avg_urgency = subgroup_df['urgency'].mean()
                                    st.metric("Avg Urgency", f"{avg_urgency:.1f}")

                                # Radar chart for subgroup profile
                                st.markdown("##### Subgroup Profile")

                                categories = ['Power', 'Interest', 'Legitimacy', 'Urgency']
                                values = [avg_power, avg_interest, avg_legitimacy, avg_urgency]

                                fig = go.Figure()

                                fig.add_trace(go.Scatterpolar(
                                    r=values,
                                    theta=categories,
                                    fill='toself',
                                    name=analysis_subgroup
                                ))

                                fig.update_layout(
                                    polar=dict(
                                        radialaxis=dict(
                                            visible=True,
                                            range=[0, 10]
                                        )),
                                    showlegend=True,
                                    title=f"{analysis_subgroup} - Average Ratings Profile"
                                )

                                st.plotly_chart(fig, use_container_width=True)

                                # Detailed member ratings
                                st.markdown("##### Member Details")
                                st.dataframe(
                                    subgroup_df[['stakeholder', 'power', 'interest', 'legitimacy', 'urgency']],
                                    use_container_width=True
                                )

                                # Strategic recommendations
                                st.markdown("##### Strategic Recommendations")

                                # Determine strategy based on average scores
                                if avg_power > 5 and avg_interest > 5:
                                    strategy = "**Manage Closely** - This is a key stakeholder group requiring active engagement and regular communication."
                                elif avg_power > 5 and avg_interest <= 5:
                                    strategy = "**Keep Satisfied** - Maintain satisfaction with this influential group while monitoring their interest levels."
                                elif avg_power <= 5 and avg_interest > 5:
                                    strategy = "**Keep Informed** - Regular updates and information sharing will maintain their support and engagement."
                                else:
                                    strategy = "**Monitor** - Basic monitoring is sufficient, but stay alert to changes in their position."

                                st.info(strategy)
                else:
                    st.warning("⚠️ Please create at least one subgroup first.")
        else:
            st.warning("⚠️ Please complete Step 1 to rate stakeholders.")
    
    with tab3:
        st.markdown('<p class="step-header">Step 3: Conflict Resolution Strategy</p>', unsafe_allow_html=True)
        st.info("Identify conflict resolution approaches based on cooperativeness vs competition.")
        
        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            st.subheader("Identify Stakeholder Conflicts")
            
            stakeholders = [item['stakeholder'] for item in st.session_state.sat_data['relationship_data']]
            
            selected_sh = st.selectbox("Select Stakeholder for Conflict Analysis", stakeholders)
            
            cooperativeness = st.slider("Cooperativeness Level", 1, 10, 5)
            competitiveness = st.slider("Competitiveness Level", 1, 10, 5)
            
            conflict_description = st.text_area("Describe potential conflicts")
            
            if st.button("Add Conflict Strategy"):
                if 'conflict_data' not in st.session_state.sat_data:
                    st.session_state.sat_data['conflict_data'] = []
                
                st.session_state.sat_data['conflict_data'].append({
                    'stakeholder': selected_sh,
                    'cooperativeness': cooperativeness,
                    'competitiveness': competitiveness,
                    'description': conflict_description
                })
                st.success("✅ Conflict strategy added!")
            
            # Show conflict matrix
            if 'conflict_data' in st.session_state.sat_data and st.session_state.sat_data['conflict_data']:
                df_conflict = pd.DataFrame(st.session_state.sat_data['conflict_data'])

                with st.expander("🗑️ Manage conflict entries"):
                    for idx, c in enumerate(list(st.session_state.sat_data['conflict_data'])):
                        cc1, cc2 = st.columns([5, 1])
                        with cc1:
                            st.write(f"**{idx+1}.** {c.get('stakeholder','?')} — "
                                     f"coop:{c.get('cooperativeness','-')}, "
                                     f"comp:{c.get('competitiveness','-')}")
                        with cc2:
                            if st.button("🗑️ Delete", key=f"del_conflict_{idx}"):
                                st.session_state.sat_data['conflict_data'].pop(idx)
                                st.rerun()

                fig = px.scatter(df_conflict, x='competitiveness', y='cooperativeness',
                               text='stakeholder',
                               title='Conflict Resolution Matrix')
                fig.update_traces(textposition='top center')
                fig.add_hline(y=5, line_dash="dash", line_color="gray")
                fig.add_vline(x=5, line_dash="dash", line_color="gray")
                st.plotly_chart(fig, use_container_width=True)
                
                st.markdown("""
                **Strategies:**
                - **High Cooperativeness, Low Competition:** Accommodation
                - **High Cooperativeness, High Competition:** Collaboration
                - **Low Cooperativeness, Low Competition:** Avoidance
                - **Low Cooperativeness, High Competition:** Competition
                - **Moderate Both:** Compromise
                """)
        else:
            st.warning("⚠️ Please complete Step 1 first.")
    
    with tab4:
        st.markdown('<p class="step-header">Step 4: Knowledge and Responsibility Chart</p>', unsafe_allow_html=True)
        st.info("Map knowledge, skills, and responsibilities for each stakeholder group.")
        
        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            st.subheader("Define Knowledge and Responsibilities")
            
            stakeholder_groups = st.multiselect(
                "Select Stakeholder Group",
                ["Key Players", "Keep Satisfied", "Keep Informed", "Monitor"],
                default=["Key Players"]
            )
            
            for group in stakeholder_groups:
                with st.expander(f"📚 {group}"):
                    knowledge_areas = st.text_area(f"Knowledge Areas for {group}", key=f"know_{group}")
                    responsibilities = st.text_area(f"Responsibilities for {group}", key=f"resp_{group}")
                    skills_needed = st.text_area(f"Skills Needed for {group}", key=f"skill_{group}")
                    
                    if st.button(f"Save {group} Data", key=f"btn_{group}"):
                        if 'knowledge_data' not in st.session_state.sat_data:
                            st.session_state.sat_data['knowledge_data'] = {}
                        
                        st.session_state.sat_data['knowledge_data'][group] = {
                            'knowledge': knowledge_areas,
                            'responsibilities': responsibilities,
                            'skills': skills_needed
                        }
                        st.success(f"✅ Data saved for {group}")
            
            # Display summary
            if 'knowledge_data' in st.session_state.sat_data:
                st.subheader("Knowledge & Responsibility Summary")
                for group in list(st.session_state.sat_data['knowledge_data'].keys()):
                    data = st.session_state.sat_data['knowledge_data'][group]
                    c1, c2 = st.columns([6, 1])
                    with c1:
                        st.markdown(f"**{group}:**")
                        st.write(f"- Knowledge: {data['knowledge']}")
                        st.write(f"- Responsibilities: {data['responsibilities']}")
                        st.write(f"- Skills: {data['skills']}")
                    with c2:
                        if st.button("🗑️", key=f"del_knowledge_{group}", help=f"Delete {group} entry"):
                            del st.session_state.sat_data['knowledge_data'][group]
                            st.rerun()
        else:
            st.warning("⚠️ Please complete Step 1 first.")
    
    with tab5:
        st.markdown('<p class="step-header">Step 5: Value Exchange Map</p>', unsafe_allow_html=True)
        st.info("Identify pains, gains, and strategies for training-related product-service systems.")
        
        if 'relationship_data' in st.session_state.sat_data and st.session_state.sat_data['relationship_data']:
            st.subheader("Value Proposition Canvas")
            
            stakeholder = st.selectbox(
                "Select Stakeholder for Value Mapping",
                [item['stakeholder'] for item in st.session_state.sat_data['relationship_data']]
            )
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### Customer Profile")
                pains = st.text_area("Pains (problems, challenges)")
                gains = st.text_area("Gains (desired outcomes, benefits)")
                jobs = st.text_area("Jobs to be Done")
            
            with col2:
                st.markdown("### Value Map")
                pain_relievers = st.text_area("Pain Relievers (how training helps)")
                gain_creators = st.text_area("Gain Creators (benefits provided)")
                products_services = st.text_area("Training Products & Services")
            
            if st.button("Save Value Map"):
                if 'value_map' not in st.session_state.sat_data:
                    st.session_state.sat_data['value_map'] = []
                
                st.session_state.sat_data['value_map'].append({
                    'stakeholder': stakeholder,
                    'pains': pains,
                    'gains': gains,
                    'jobs': jobs,
                    'pain_relievers': pain_relievers,
                    'gain_creators': gain_creators,
                    'products_services': products_services
                })
                st.success("✅ Value map saved!")
            
            # Display existing value maps
            if 'value_map' in st.session_state.sat_data and st.session_state.sat_data['value_map']:
                st.subheader("Saved Value Maps")
                for vm_idx, vm in enumerate(list(st.session_state.sat_data['value_map'])):
                    with st.expander(f"📊 {vm['stakeholder']}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("**Customer Profile:**")
                            st.write(f"Pains: {vm['pains']}")
                            st.write(f"Gains: {vm['gains']}")
                            st.write(f"Jobs: {vm['jobs']}")
                        with col2:
                            st.markdown("**Value Map:**")
                            st.write(f"Pain Relievers: {vm['pain_relievers']}")
                            st.write(f"Gain Creators: {vm['gain_creators']}")
                            st.write(f"Products/Services: {vm['products_services']}")
                        if st.button("🗑️ Delete this value map", key=f"del_vm_{vm_idx}"):
                            st.session_state.sat_data['value_map'].pop(vm_idx)
                            st.rerun()
        else:
            st.warning("⚠️ Please complete Step 1 first.")

# MAT - Market Analysis Toolkit
elif menu == "3️⃣ MAT - Market Analysis":
    st.markdown('<div class="main-header">MAT - Market Analysis Toolkit</div>', unsafe_allow_html=True)
    
    st.markdown("""
    MAT ensures that training programs align with market demands and opportunities through comprehensive market analysis.
    """)
    
    mat_tools = st.selectbox("Select Analysis Tool", [
        "PESTEL Analysis",
        "Gap Analysis",
        "Behavioral Segmentation",
        "User Persona",
        "Customer Journey Map",
        "Mystery Shopping",
        "Complaint Data Analysis",
        "Brand Audit"
    ])

    # Map the tool label to its mat_data key(s) so users can clear a whole tool's data.
    _mat_tool_keys = {
        "PESTEL Analysis": ["pestel"],
        "Gap Analysis": ["gap"],
        "Behavioral Segmentation": ["behavioral_segments"],
        "User Persona": ["personas"],
        "Customer Journey Map": ["customer_journey"],
        "Mystery Shopping": ["mystery_shopping"],
        "Complaint Data Analysis": ["complaints"],
        "Brand Audit": ["brand_audit"],
    }
    _clear_keys = [k for k in _mat_tool_keys.get(mat_tools, []) if k in st.session_state.mat_data]
    if _clear_keys:
        with st.expander(f"🗑️ Clear all {mat_tools} data"):
            st.caption("This removes every entry recorded under this tool. Cannot be undone (unless you saved the project).")
            if st.button(f"🗑️ Clear {mat_tools} data", key=f"_clear_mat_{mat_tools}"):
                for k in _clear_keys:
                    st.session_state.mat_data.pop(k, None)
                st.success(f"Cleared {mat_tools} data.")
                st.rerun()

    if mat_tools == "PESTEL Analysis":
        st.subheader("PESTEL Analysis")
        st.info("Analyze Political, Economic, Social, Technological, Environmental, and Legal factors.")
        
        col1, col2 = st.columns(2)
        
        with col1:
            political = st.text_area(
                "Political Factors",
                placeholder="e.g., PM Vishwakarma scheme offers 15% subsidy on raw materials"
            )
            economic = st.text_area(
                "Economic Factors",
                placeholder="e.g., Rising demand for handmade products in urban markets"
            )
            social = st.text_area(
                "Social Factors",
                placeholder="e.g., Growing appreciation for traditional crafts among youth"
            )
        
        with col2:
            technological = st.text_area(
                "Technological Factors",
                placeholder="e.g., E-commerce platforms enabling direct sales"
            )
            environmental = st.text_area(
                "Environmental Factors",
                placeholder="e.g., Preference for sustainable, eco-friendly products"
            )
            legal = st.text_area(
                "Legal Factors",
                placeholder="e.g., GI tag protection for regional crafts"
            )
        
        if st.button("Save PESTEL Analysis"):
            st.session_state.mat_data['pestel'] = {
                'political': political,
                'economic': economic,
                'social': social,
                'technological': technological,
                'environmental': environmental,
                'legal': legal,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            st.success("✅ PESTEL Analysis saved!")
    
    elif mat_tools == "Gap Analysis":
        st.subheader("Gap Analysis")
        st.info("Identify the gap between current state and desired future state.")
        
        col1, col2 = st.columns(2)
        
        with col1:
            current_state = st.text_area(
                "Current State",
                placeholder="e.g., Market for handmade wedding invitations sees very seasonal demand"
            )
            current_strengths = st.text_area("Current Strengths")
            current_weaknesses = st.text_area("Current Weaknesses")
        
        with col2:
            desired_state = st.text_area(
                "Desired Future State",
                placeholder="e.g., Year-round consistent demand with diversified product line"
            )
            opportunities = st.text_area("Opportunities")
            threats = st.text_area("Threats")
        
        action_plan = st.text_area("Action Plan to Bridge the Gap")
        
        if st.button("Save Gap Analysis"):
            st.session_state.mat_data['gap'] = {
                'current_state': current_state,
                'current_strengths': current_strengths,
                'current_weaknesses': current_weaknesses,
                'desired_state': desired_state,
                'opportunities': opportunities,
                'threats': threats,
                'action_plan': action_plan,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            st.success("✅ Gap Analysis saved!")
    
    elif mat_tools == "Behavioral Segmentation":
        st.subheader("Behavioral Segmentation")
        st.info("Segment customers based on purchasing behavior, usage, and benefits sought.")
        
        segment_name = st.text_input("Segment Name")
        
        col1, col2 = st.columns(2)
        
        with col1:
            purchase_behavior = st.text_area("Purchase Behavior")
            usage_rate = st.text_area("Usage Rate")
        
        with col2:
            benefits_sought = st.text_area("Benefits Sought")
            loyalty_status = st.text_area("Loyalty Status")
        
        occasion = st.text_area("Purchase Occasion")
        
        if st.button("Add Segment"):
            if 'behavioral_segments' not in st.session_state.mat_data:
                st.session_state.mat_data['behavioral_segments'] = []
            
            st.session_state.mat_data['behavioral_segments'].append({
                'name': segment_name,
                'purchase_behavior': purchase_behavior,
                'usage_rate': usage_rate,
                'benefits_sought': benefits_sought,
                'loyalty_status': loyalty_status,
                'occasion': occasion
            })
            st.success(f"✅ Segment '{segment_name}' added!")
        
        # Display segments
        if 'behavioral_segments' in st.session_state.mat_data:
            st.subheader("Customer Segments")
            for seg_idx, seg in enumerate(list(st.session_state.mat_data['behavioral_segments'])):
                with st.expander(f"👥 {seg['name']}"):
                    st.write(f"**Purchase Behavior:** {seg['purchase_behavior']}")
                    st.write(f"**Usage Rate:** {seg['usage_rate']}")
                    st.write(f"**Benefits Sought:** {seg['benefits_sought']}")
                    st.write(f"**Loyalty Status:** {seg['loyalty_status']}")
                    st.write(f"**Occasion:** {seg['occasion']}")
                    if st.button("🗑️ Delete this segment", key=f"del_seg_{seg_idx}"):
                        st.session_state.mat_data['behavioral_segments'].pop(seg_idx)
                        st.rerun()
    
    elif mat_tools == "User Persona":
        st.subheader("User Persona")
        st.info("Create detailed profiles of typical customers.")
        
        persona_name = st.text_input("Persona Name")
        
        col1, col2 = st.columns(2)
        
        with col1:
            age_range = st.text_input("Age Range")
            occupation = st.text_input("Occupation")
            location = st.text_input("Location")
            income_level = st.text_input("Income Level")
        
        with col2:
            education = st.text_input("Education Level")
            family_status = st.text_input("Family Status")
            lifestyle = st.text_input("Lifestyle")
        
        goals = st.text_area("Goals & Motivations")
        pain_points = st.text_area("Pain Points & Frustrations")
        shopping_habits = st.text_area("Shopping Habits")
        
        if st.button("Save Persona"):
            if 'personas' not in st.session_state.mat_data:
                st.session_state.mat_data['personas'] = []
            
            st.session_state.mat_data['personas'].append({
                'name': persona_name,
                'age_range': age_range,
                'occupation': occupation,
                'location': location,
                'income_level': income_level,
                'education': education,
                'family_status': family_status,
                'lifestyle': lifestyle,
                'goals': goals,
                'pain_points': pain_points,
                'shopping_habits': shopping_habits
            })
            st.success(f"✅ Persona '{persona_name}' created!")
        
        # Display personas
        if 'personas' in st.session_state.mat_data:
            st.subheader("Created Personas")
            for p_idx, persona in enumerate(list(st.session_state.mat_data['personas'])):
                with st.expander(f"👤 {persona['name']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.write(f"**Demographics:**")
                        st.write(f"- Age: {persona['age_range']}")
                        st.write(f"- Occupation: {persona['occupation']}")
                        st.write(f"- Location: {persona['location']}")
                        st.write(f"- Income: {persona['income_level']}")
                    with col2:
                        st.write(f"**Psychographics:**")
                        st.write(f"- Education: {persona['education']}")
                        st.write(f"- Family: {persona['family_status']}")
                        st.write(f"- Lifestyle: {persona['lifestyle']}")
                    st.write(f"**Goals:** {persona['goals']}")
                    st.write(f"**Pain Points:** {persona['pain_points']}")
                    st.write(f"**Shopping Habits:** {persona['shopping_habits']}")
                    if st.button("🗑️ Delete this persona", key=f"del_persona_{p_idx}"):
                        st.session_state.mat_data['personas'].pop(p_idx)
                        st.rerun()
    
    elif mat_tools == "Customer Journey Map":
        st.subheader("Customer Journey Map")
        st.info("Map the customer's experience from awareness to post-purchase.")
        
        stages = ["Awareness", "Consideration", "Purchase", "Usage", "Loyalty"]
        
        journey_data = {}
        for stage in stages:
            with st.expander(f"📍 {stage}"):
                journey_data[stage] = {
                    'touchpoints': st.text_area(f"Touchpoints in {stage}", key=f"tp_{stage}"),
                    'actions': st.text_area(f"Customer Actions", key=f"act_{stage}"),
                    'emotions': st.text_area(f"Emotions/Thoughts", key=f"emo_{stage}"),
                    'pain_points': st.text_area(f"Pain Points", key=f"pain_{stage}"),
                    'opportunities': st.text_area(f"Opportunities for Improvement", key=f"opp_{stage}")
                }
        
        if st.button("Save Customer Journey"):
            st.session_state.mat_data['customer_journey'] = journey_data
            st.success("✅ Customer Journey Map saved!")
    
    elif mat_tools == "Mystery Shopping":
        st.subheader("Mystery Shopping")
        st.info("Evaluate customer experience by acting as a customer.")
        
        location = st.text_input("Location Evaluated")
        date = st.date_input("Date of Visit")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("**Store/Online Experience:**")
            ambiance = st.slider("Ambiance/Interface", 1, 10, 5)
            accessibility = st.slider("Accessibility", 1, 10, 5)
            product_display = st.slider("Product Display", 1, 10, 5)
        
        with col2:
            st.markdown("**Service Quality:**")
            staff_behavior = st.slider("Staff Behavior", 1, 10, 5)
            product_knowledge = st.slider("Product Knowledge", 1, 10, 5)
            response_time = st.slider("Response Time", 1, 10, 5)
        
        observations = st.text_area("Detailed Observations")
        recommendations = st.text_area("Recommendations for Improvement")
        
        if st.button("Save Mystery Shopping Report"):
            if 'mystery_shopping' not in st.session_state.mat_data:
                st.session_state.mat_data['mystery_shopping'] = []
            
            st.session_state.mat_data['mystery_shopping'].append({
                'location': location,
                'date': str(date),
                'ambiance': ambiance,
                'accessibility': accessibility,
                'product_display': product_display,
                'staff_behavior': staff_behavior,
                'product_knowledge': product_knowledge,
                'response_time': response_time,
                'observations': observations,
                'recommendations': recommendations
            })
            st.success("✅ Mystery Shopping Report saved!")

        if 'mystery_shopping' in st.session_state.mat_data and st.session_state.mat_data['mystery_shopping']:
            st.subheader("Saved Mystery Shopping Reports")
            for ms_idx, ms in enumerate(list(st.session_state.mat_data['mystery_shopping'])):
                with st.expander(f"🛍️ {ms.get('location','?')} — {ms.get('date','')}"):
                    st.write(f"Ambiance: {ms.get('ambiance')}, Accessibility: {ms.get('accessibility')}, "
                             f"Display: {ms.get('product_display')}, Staff: {ms.get('staff_behavior')}, "
                             f"Knowledge: {ms.get('product_knowledge')}, Response: {ms.get('response_time')}")
                    if ms.get('observations'):
                        st.write(f"**Observations:** {ms['observations']}")
                    if ms.get('recommendations'):
                        st.write(f"**Recommendations:** {ms['recommendations']}")
                    if st.button("🗑️ Delete this report", key=f"del_ms_{ms_idx}"):
                        st.session_state.mat_data['mystery_shopping'].pop(ms_idx)
                        st.rerun()

    elif mat_tools == "Complaint Data Analysis":
        st.subheader("Complaint Data Analysis")
        st.info("Analyze customer complaints to identify patterns and improvement areas.")
        
        complaint_source = st.text_input("Complaint Source")
        complaint_date = st.date_input("Date Received")
        
        complaint_category = st.selectbox(
            "Complaint Category",
            ["Product Quality", "Service", "Delivery", "Pricing", "Communication", "Other"]
        )
        
        complaint_description = st.text_area("Complaint Description")
        severity = st.select_slider("Severity", options=["Low", "Medium", "High", "Critical"])
        
        resolution_status = st.selectbox("Resolution Status", 
                                        ["Pending", "In Progress", "Resolved", "Closed"])
        resolution_details = st.text_area("Resolution Details")
        
        if st.button("Add Complaint"):
            if 'complaints' not in st.session_state.mat_data:
                st.session_state.mat_data['complaints'] = []
            
            st.session_state.mat_data['complaints'].append({
                'source': complaint_source,
                'date': str(complaint_date),
                'category': complaint_category,
                'description': complaint_description,
                'severity': severity,
                'status': resolution_status,
                'resolution': resolution_details
            })
            st.success("✅ Complaint logged!")
        
        # Analysis
        if 'complaints' in st.session_state.mat_data and st.session_state.mat_data['complaints']:
            st.subheader("Complaint Analysis")
            
            df_complaints = pd.DataFrame(st.session_state.mat_data['complaints'])
            
            col1, col2 = st.columns(2)
            
            with col1:
                # Category distribution
                category_counts = df_complaints['category'].value_counts()
                fig_cat = px.pie(values=category_counts.values, 
                               names=category_counts.index,
                               title="Complaints by Category")
                st.plotly_chart(fig_cat, use_container_width=True)
            
            with col2:
                # Severity distribution
                severity_counts = df_complaints['severity'].value_counts()
                fig_sev = px.bar(x=severity_counts.index,
                               y=severity_counts.values,
                               title="Complaints by Severity")
                st.plotly_chart(fig_sev, use_container_width=True)

            with st.expander("🗑️ Manage complaint entries"):
                for idx, c in enumerate(list(st.session_state.mat_data['complaints'])):
                    cc1, cc2 = st.columns([5, 1])
                    with cc1:
                        st.write(f"**{idx+1}.** [{c.get('severity','?')}] {c.get('category','?')} — "
                                 f"{c.get('source','?')} ({c.get('date','')})")
                        if c.get('description'):
                            st.caption(c['description'][:120])
                    with cc2:
                        if st.button("🗑️ Delete", key=f"del_complaint_{idx}"):
                            st.session_state.mat_data['complaints'].pop(idx)
                            st.rerun()

    elif mat_tools == "Brand Audit":
        st.subheader("Brand Audit")
        st.info("Evaluate brand perception, positioning, and performance.")
        
        st.markdown("### Brand Identity")
        col1, col2 = st.columns(2)
        
        with col1:
            brand_mission = st.text_area("Brand Mission")
            brand_vision = st.text_area("Brand Vision")
            brand_values = st.text_area("Brand Values")
        
        with col2:
            unique_selling_proposition = st.text_area("Unique Selling Proposition")
            brand_personality = st.text_area("Brand Personality")
            brand_promise = st.text_area("Brand Promise")
        
        st.markdown("### Brand Performance")
        col1, col2, col3 = st.columns(3)
        
        with col1:
            brand_awareness = st.slider("Brand Awareness", 1, 10, 5)
            brand_recognition = st.slider("Brand Recognition", 1, 10, 5)
        
        with col2:
            brand_loyalty = st.slider("Brand Loyalty", 1, 10, 5)
            customer_satisfaction = st.slider("Customer Satisfaction", 1, 10, 5)
        
        with col3:
            market_position = st.slider("Market Position", 1, 10, 5)
            brand_consistency = st.slider("Brand Consistency", 1, 10, 5)
        
        strengths = st.text_area("Brand Strengths")
        weaknesses = st.text_area("Brand Weaknesses")
        recommendations = st.text_area("Strategic Recommendations")
        
        if st.button("Save Brand Audit"):
            st.session_state.mat_data['brand_audit'] = {
                'brand_mission': brand_mission,
                'brand_vision': brand_vision,
                'brand_values': brand_values,
                'usp': unique_selling_proposition,
                'personality': brand_personality,
                'promise': brand_promise,
                'awareness': brand_awareness,
                'recognition': brand_recognition,
                'loyalty': brand_loyalty,
                'satisfaction': customer_satisfaction,
                'position': market_position,
                'consistency': brand_consistency,
                'strengths': strengths,
                'weaknesses': weaknesses,
                'recommendations': recommendations,
                'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            st.success("✅ Brand Audit saved!")

# Nature of Craft
elif menu == "4️⃣ Nature of Craft":
    st.markdown('<div class="main-header">Nature of Craft - 5P Framework</div>', unsafe_allow_html=True)

    st.markdown("""
    The 5P Framework helps define and analyze the nature of your craft across five key dimensions:
    **Product**, **Proficiency**, **Process**, **Purpose**, and **Portrayal**.
    """)

    # Display schematic diagram
    st.markdown("---")
    st.subheader("📐 Schematic Diagram of Craft Definition")

    col1, col2 = st.columns([2, 1])

    with col1:
        st.image("schematic.jpg", caption="Figure 2.1: Schematic diagram of definition of craft", use_column_width=True)

    with col2:
        st.markdown("""
        ### The 5P Layers

        From innermost to outermost:

        1. **Portrayal** (Core)
        2. **Purpose**
        3. **Process**
        4. **Proficiency**
        5. **Product** (Outer)

        Each layer represents a different aspect of craft nature, from its deepest meaning to its market manifestation.
        """)

    st.markdown("---")

    # Initialize session state for nature of craft
    if 'nature_of_craft' not in st.session_state:
        st.session_state.nature_of_craft = {
            'current_status': {},
            'desired_status': {}
        }

    # Define the 5P Framework data structure
    framework_data = {
        "Product (Any Craft product)": {
            "secondary": ["Utilitarian", "Decorative Artefacts"],
            "tertiary": ["Functional Utility"]
        },
        "Proficiency (Needs a skill)": {
            "secondary": ["Material Understanding", "Practical knowledge", "Craftsmanship", "Experience", "Emotional Value Creation"],
            "tertiary": ["Skill-based Activity", "Intuitive Learning", "Skill sharing", "Craft Disciplines", "Sensory Experience", "Aesthetic Judgement", "Craft Knowledge"]
        },
        "Process (follows a process)": {
            "secondary": ["Sustainable manufacturing", "Material Manipulation Techniques", "Physical World Interaction", "Local Production", "Network Engagement"],
            "tertiary": ["Eco-effective process", "Non-industrial Production", "Community Collaboration method"]
        },
        "Purpose (to fulfil a purpose)": {
            "secondary": ["Prosumption", "Contextual lifestyle", "Economical aspect", "Self-satisfaction", "Aesthetic value"],
            "tertiary": ["Consumer Market Focus", "Consumer Behaviour", "Community Economy"]
        },
        "Portrayal (portrays a meaning)": {
            "secondary": ["Social Significance", "Creative Expression", "Cultural and Religious Representation"],
            "tertiary": ["Traditional Folkloric", "Ideology", "Local Culture", "Cultural Heritage", "Cultural Symbolism", "Traditional Wisdom", "Self-Expression", "Individual Works Conception"]
        }
    }

    # Create tabs for Current and Desired Status
    status_tabs = st.tabs(["📍 Current Status", "🎯 Desired New Status"])

    # Tab 1: Current Status
    with status_tabs[0]:
        st.markdown("### Current Status of Your Craft")
        st.info("Select the characteristics that best describe your craft's **current** nature across the 5P dimensions.")

        for primary, content in framework_data.items():
            with st.expander(f"**{primary}**", expanded=False):
                st.markdown(f"##### {primary}")

                # Secondary Nature
                st.markdown("**Secondary Nature:**")
                for item in content['secondary']:
                    key = f"current_{primary}_{item}"
                    if key not in st.session_state.nature_of_craft['current_status']:
                        st.session_state.nature_of_craft['current_status'][key] = False

                    st.session_state.nature_of_craft['current_status'][key] = st.checkbox(
                        item,
                        value=st.session_state.nature_of_craft['current_status'][key],
                        key=f"cb_{key}"
                    )

                # Tertiary Nature (if exists)
                if content['tertiary']:
                    st.markdown("**Tertiary Nature:**")
                    for item in content['tertiary']:
                        key = f"current_{primary}_{item}"
                        if key not in st.session_state.nature_of_craft['current_status']:
                            st.session_state.nature_of_craft['current_status'][key] = False

                        st.session_state.nature_of_craft['current_status'][key] = st.checkbox(
                            item,
                            value=st.session_state.nature_of_craft['current_status'][key],
                            key=f"cb_{key}"
                        )

        # Save Current Status
        if st.button("💾 Save Current Status", use_container_width=True, type="primary"):
            # Count selections
            current_selections = sum(1 for v in st.session_state.nature_of_craft['current_status'].values() if v)
            st.success(f"✅ Current status saved! ({current_selections} characteristics selected)")

    # Tab 2: Desired New Status
    with status_tabs[1]:
        st.markdown("### Desired New Status of Your Craft")
        st.info("Select the characteristics you want your craft to **achieve** or **develop** across the 5P dimensions.")

        for primary, content in framework_data.items():
            with st.expander(f"**{primary}**", expanded=False):
                st.markdown(f"##### {primary}")

                # Secondary Nature
                st.markdown("**Secondary Nature:**")
                for item in content['secondary']:
                    key = f"desired_{primary}_{item}"
                    if key not in st.session_state.nature_of_craft['desired_status']:
                        st.session_state.nature_of_craft['desired_status'][key] = False

                    st.session_state.nature_of_craft['desired_status'][key] = st.checkbox(
                        item,
                        value=st.session_state.nature_of_craft['desired_status'][key],
                        key=f"cb_{key}"
                    )

                # Tertiary Nature (if exists)
                if content['tertiary']:
                    st.markdown("**Tertiary Nature:**")
                    for item in content['tertiary']:
                        key = f"desired_{primary}_{item}"
                        if key not in st.session_state.nature_of_craft['desired_status']:
                            st.session_state.nature_of_craft['desired_status'][key] = False

                        st.session_state.nature_of_craft['desired_status'][key] = st.checkbox(
                            item,
                            value=st.session_state.nature_of_craft['desired_status'][key],
                            key=f"cb_{key}"
                        )

        # Save Desired Status
        if st.button("💾 Save Desired Status", use_container_width=True, type="primary"):
            # Count selections
            desired_selections = sum(1 for v in st.session_state.nature_of_craft['desired_status'].values() if v)
            st.success(f"✅ Desired status saved! ({desired_selections} characteristics selected)")

    # Analysis Section
    st.markdown("---")
    st.subheader("📊 Gap Analysis")

    if any(st.session_state.nature_of_craft['current_status'].values()) or any(st.session_state.nature_of_craft['desired_status'].values()):

        # Calculate statistics for each P
        analysis_data = []

        for primary in framework_data.keys():
            # Count current selections for this P
            current_count = sum(
                1 for k, v in st.session_state.nature_of_craft['current_status'].items()
                if v and k.startswith(f"current_{primary}")
            )

            # Count desired selections for this P
            desired_count = sum(
                1 for k, v in st.session_state.nature_of_craft['desired_status'].items()
                if v and k.startswith(f"desired_{primary}")
            )

            # Calculate gap
            gap = desired_count - current_count

            analysis_data.append({
                "Dimension": primary.split("(")[0].strip(),
                "Current": current_count,
                "Desired": desired_count,
                "Gap": gap,
                "Status": "✅ Achieved" if gap <= 0 else f"📈 Need {gap} more"
            })

        df_analysis = pd.DataFrame(analysis_data)
        st.dataframe(df_analysis, use_container_width=True)

        # Visualization
        col1, col2 = st.columns(2)

        with col1:
            # Bar chart comparing current vs desired
            fig = go.Figure()

            fig.add_trace(go.Bar(
                name='Current',
                x=df_analysis['Dimension'],
                y=df_analysis['Current'],
                marker_color='lightblue'
            ))

            fig.add_trace(go.Bar(
                name='Desired',
                x=df_analysis['Dimension'],
                y=df_analysis['Desired'],
                marker_color='darkgreen'
            ))

            fig.update_layout(
                title='Current vs Desired Status by Dimension',
                barmode='group',
                xaxis_title='5P Dimensions',
                yaxis_title='Number of Characteristics',
                height=400
            )

            st.plotly_chart(fig, use_container_width=True)

        with col2:
            # Gap visualization
            fig2 = go.Figure()

            colors = ['green' if x <= 0 else 'orange' for x in df_analysis['Gap']]

            fig2.add_trace(go.Bar(
                x=df_analysis['Dimension'],
                y=df_analysis['Gap'],
                marker_color=colors,
                text=df_analysis['Gap'],
                textposition='auto'
            ))

            fig2.update_layout(
                title='Gap Analysis (Desired - Current)',
                xaxis_title='5P Dimensions',
                yaxis_title='Gap (positive = need development)',
                height=400
            )

            st.plotly_chart(fig2, use_container_width=True)

        # Recommendations
        st.markdown("### 💡 Development Recommendations")

        for _, row in df_analysis.iterrows():
            if row['Gap'] > 0:
                st.warning(f"**{row['Dimension']}**: Focus on developing {row['Gap']} additional characteristic(s) to reach your desired state.")
            elif row['Gap'] == 0 and row['Current'] > 0:
                st.success(f"**{row['Dimension']}**: You've achieved your desired state in this dimension!")

        # Detailed comparison
        st.markdown("---")
        st.markdown("### 📋 Detailed Status Comparison")

        comparison_data = []

        for primary in framework_data.keys():
            # Get all items for this primary
            all_items = framework_data[primary]['secondary'] + framework_data[primary]['tertiary']

            for item in all_items:
                current_key = f"current_{primary}_{item}"
                desired_key = f"desired_{primary}_{item}"

                current_val = st.session_state.nature_of_craft['current_status'].get(current_key, False)
                desired_val = st.session_state.nature_of_craft['desired_status'].get(desired_key, False)

                if current_val or desired_val:  # Only show if selected in either
                    status = ""
                    if current_val and desired_val:
                        status = "✅ Maintained"
                    elif current_val and not desired_val:
                        status = "⬇️ To reduce"
                    elif not current_val and desired_val:
                        status = "⬆️ To develop"

                    comparison_data.append({
                        "Dimension": primary.split("(")[0].strip(),
                        "Characteristic": item,
                        "Current": "✓" if current_val else "✗",
                        "Desired": "✓" if desired_val else "✗",
                        "Action": status
                    })

        if comparison_data:
            df_comparison = pd.DataFrame(comparison_data)
            st.dataframe(df_comparison, use_container_width=True)
        else:
            st.info("Select characteristics in both Current and Desired status to see detailed comparison.")

    else:
        st.info("👆 Complete the Current and Desired Status sections above to see gap analysis and recommendations.")

# Projects
elif menu == "📁 Projects":
    st.markdown('<div class="main-header">Projects</div>', unsafe_allow_html=True)

    if st.session_state.pop("_loaded_from_browser", None):
        st.success(f"✅ Loaded {len(st.session_state.projects)} project(s) from browser storage.")
    if st.session_state.pop("_load_error", None):
        st.error(f"Browser load error: {st.session_state.get('_load_error', '')}")

    st.markdown("""
        <div class="toolkit-card">
        <h3>📁 Manage Projects</h3>
        Save, load, rename, and delete SIAMA analyses as named projects. Each project bundles your
        SIT, SAT, MAT, and Nature of Craft data. Use <b>Browser Storage</b> to keep projects on this
        device between visits, and <b>Export/Import</b> to move them to other devices or share with others.
        </div>
    """, unsafe_allow_html=True)

    if st.session_state.current_project:
        st.info(f"🟢 Currently editing: **{st.session_state.current_project}**")
    else:
        st.info("⚪ No project loaded. Your current edits are held in this session only — save them as a project to keep them.")

    st.markdown("---")

    tab_list, tab_create, tab_transfer = st.tabs(
        ["📋 My Projects", "➕ New / Save Current", "💾 Browser / Export / Import"]
    )

    with tab_create:
        st.markdown("### Save current work as a new project")
        new_name = st.text_input("Project name", key="_new_project_name", placeholder="e.g. Bandhani Training Program – Kutch 2026")
        cc1, cc2 = st.columns(2)
        with cc1:
            if st.button("💾 Save as new project", type="primary", key="_save_new_btn"):
                name = (new_name or "").strip()
                if not name:
                    st.error("Project name cannot be empty.")
                elif name in st.session_state.projects:
                    st.error(f"A project named '{name}' already exists. Rename it, or update it from the '📋 My Projects' tab.")
                else:
                    now = datetime.now().isoformat(timespec="seconds")
                    st.session_state.projects[name] = {
                        "created": now, "updated": now, **snapshot_current_state()
                    }
                    st.session_state.current_project = name
                    st.success(f"✅ Saved project '{name}'.")
                    st.rerun()
        with cc2:
            if st.button("🆕 Start new blank project", key="_new_blank_btn"):
                reset_state()
                st.session_state.current_project = None
                st.success("Started a new blank analysis. Current data has been cleared.")
                st.rerun()

        if st.session_state.current_project:
            st.markdown("---")
            st.markdown(f"### Update the currently loaded project")
            if st.button(f"💾 Save changes to '{st.session_state.current_project}'", type="primary", key="_save_update_btn"):
                name = st.session_state.current_project
                prev = st.session_state.projects.get(name, {})
                st.session_state.projects[name] = {
                    "created": prev.get("created", datetime.now().isoformat(timespec="seconds")),
                    "updated": datetime.now().isoformat(timespec="seconds"),
                    **snapshot_current_state()
                }
                st.success(f"✅ Updated '{name}'.")
                st.rerun()

    with tab_list:
        if not st.session_state.projects:
            st.info("No projects yet. Save one in the '➕ New / Save Current' tab, or load from Browser Storage / Import.")
        else:
            st.markdown(f"### {len(st.session_state.projects)} project(s)")
            for pname in sorted(st.session_state.projects.keys()):
                proj = st.session_state.projects[pname]
                header_mark = "🟢 " if pname == st.session_state.current_project else ""
                with st.expander(
                    f"{header_mark}📁 **{pname}** — updated {proj.get('updated', 'N/A')}",
                    expanded=(pname == st.session_state.current_project),
                ):
                    sit_count = len(proj.get('sit_data', {}).get('stakeholders', []))
                    sat_items = len(proj.get('sat_data', {}).get('relationship_data', []) or [])
                    mat_items = len(proj.get('mat_data', {}) or {})
                    st.caption(
                        f"SIT: {sit_count} stakeholders · SAT: {sat_items} relationships · "
                        f"MAT: {mat_items} tools · created {proj.get('created', 'N/A')}"
                    )

                    lc, rc, dc = st.columns([1, 2, 1])
                    with lc:
                        if st.button("📂 Load", key=f"load_{pname}"):
                            apply_state(proj)
                            st.session_state.current_project = pname
                            st.success(f"Loaded '{pname}'.")
                            st.rerun()
                    with rc:
                        new_n = st.text_input("Rename to", value=pname, key=f"rename_{pname}", label_visibility="collapsed")
                        if st.button("✏️ Rename", key=f"rename_btn_{pname}"):
                            target = (new_n or "").strip()
                            if not target or target == pname:
                                st.info("Enter a different name to rename.")
                            elif target in st.session_state.projects:
                                st.error(f"A project named '{target}' already exists.")
                            else:
                                st.session_state.projects[target] = st.session_state.projects.pop(pname)
                                if st.session_state.current_project == pname:
                                    st.session_state.current_project = target
                                st.success(f"Renamed to '{target}'.")
                                st.rerun()
                    with dc:
                        confirm = st.checkbox("Confirm", key=f"confirm_del_{pname}", help="Required before deleting")
                        if st.button("🗑️ Delete", key=f"del_{pname}", disabled=not confirm):
                            del st.session_state.projects[pname]
                            if st.session_state.current_project == pname:
                                st.session_state.current_project = None
                            st.success(f"Deleted '{pname}'.")
                            st.rerun()

    with tab_transfer:
        st.markdown("### 💾 Browser storage")
        st.caption("Projects are saved to this browser on this device. Clearing browser data will remove them.")
        bc1, bc2 = st.columns(2)
        with bc1:
            if st.button("💾 Save ALL projects to browser", type="primary", key="_push_browser_btn"):
                push_to_browser()
        with bc2:
            if st.button("📥 Load projects from browser", key="_pull_browser_btn"):
                pull_from_browser()

        st.markdown("---")
        st.markdown("### 📤 Export / 📥 Import (portable JSON)")

        if st.session_state.projects:
            bundle = serialize_projects(st.session_state.projects)
            st.download_button(
                "📥 Download all projects as JSON",
                data=bundle,
                file_name=f"siama_projects_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json",
                key="_download_bundle",
            )
        else:
            st.caption("No projects to export yet.")

        uploaded = st.file_uploader("Import projects JSON", type=["json"], key="_import_json")
        if uploaded is not None:
            try:
                content = uploaded.read().decode("utf-8")
                imported = deserialize_projects(content)
                if not isinstance(imported, dict):
                    st.error("Invalid format: expected a JSON object mapping project names to data.")
                else:
                    st.markdown(f"**{len(imported)} project(s)** in this file: {', '.join(list(imported.keys())[:5])}{'…' if len(imported) > 5 else ''}")
                    merge_choice = st.radio(
                        "If names collide:",
                        ["Skip duplicates", "Overwrite existing"],
                        horizontal=True,
                        key="_merge_choice",
                    )
                    if st.button("✅ Import into this session", key="_import_btn"):
                        added, overwritten, skipped = 0, 0, 0
                        for name, proj in imported.items():
                            if name in st.session_state.projects:
                                if merge_choice == "Overwrite existing":
                                    st.session_state.projects[name] = proj
                                    overwritten += 1
                                else:
                                    skipped += 1
                            else:
                                st.session_state.projects[name] = proj
                                added += 1
                        st.success(f"Imported — added: {added}, overwritten: {overwritten}, skipped: {skipped}")
                        st.rerun()
            except Exception as e:
                st.error(f"Import failed: {e}")

# Summary & Export
elif menu == "📊 Summary & Export":
    st.markdown('<div class="main-header">Summary & Export</div>', unsafe_allow_html=True)
    
    st.subheader("📋 Data Collection Summary")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("SIT - Stakeholders Identified", 
                 len(st.session_state.sit_data.get('stakeholders', [])))
        if st.session_state.sit_data.get('roles'):
            total_actors = sum(len(actors) for actors in st.session_state.sit_data['roles'].values())
            st.metric("Total Actors Mapped", total_actors)
    
    with col2:
        sat_metrics = 0
        if 'relationship_data' in st.session_state.sat_data:
            sat_metrics = len(st.session_state.sat_data['relationship_data'])
        st.metric("SAT - Stakeholders Analyzed", sat_metrics)
        
        value_maps = 0
        if 'value_map' in st.session_state.sat_data:
            value_maps = len(st.session_state.sat_data['value_map'])
        st.metric("Value Maps Created", value_maps)
    
    with col3:
        mat_tools_completed = len(st.session_state.mat_data)
        st.metric("MAT - Tools Completed", mat_tools_completed)
    
    st.markdown("---")
    
    # Export options
    st.subheader("📥 Export Data")
    
    export_format = st.radio("Select Export Format", ["JSON", "Excel"])
    
    if st.button("Generate Export File"):
        if export_format == "JSON":
            # Prepare data for JSON export
            export_data = {
                'sit_data': st.session_state.sit_data,
                'sat_data': st.session_state.sat_data,
                'mat_data': st.session_state.mat_data,
                'export_date': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
            
            # Convert to JSON string
            json_str = json.dumps(export_data, indent=2, default=str)
            
            st.download_button(
                label="📥 Download JSON",
                data=json_str,
                file_name=f"siama_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                mime="application/json"
            )
        
        elif export_format == "Excel":
            # Create Excel file with multiple sheets
            try:
                import io
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                
                wb = Workbook()
                
                # SIT Sheet
                ws_sit = wb.active
                ws_sit.title = "SIT Data"
                ws_sit['A1'] = "Stakeholder Identification Data"
                ws_sit['A1'].font = Font(bold=True, size=14)
                
                row = 3
                ws_sit[f'A{row}'] = "Role"
                ws_sit[f'B{row}'] = "Actor Name"
                ws_sit[f'C{row}'] = "Location"
                ws_sit[f'D{row}'] = "Contact"
                
                row += 1
                for role, actors in st.session_state.sit_data.get('roles', {}).items():
                    for actor in actors:
                        ws_sit[f'A{row}'] = role
                        ws_sit[f'B{row}'] = actor.get('name', '')
                        ws_sit[f'C{row}'] = actor.get('location', '')
                        ws_sit[f'D{row}'] = actor.get('contact', '')
                        row += 1
                
                # SAT Sheet
                if 'relationship_data' in st.session_state.sat_data:
                    ws_sat = wb.create_sheet("SAT Data")
                    ws_sat['A1'] = "Stakeholder Analysis Data"
                    ws_sat['A1'].font = Font(bold=True, size=14)
                    
                    df_sat = pd.DataFrame(st.session_state.sat_data['relationship_data'])
                    
                    for r_idx, row in enumerate(df_sat.values, start=3):
                        for c_idx, value in enumerate(row, start=1):
                            ws_sat.cell(row=r_idx, column=c_idx, value=str(value))
                
                # MAT Sheet
                ws_mat = wb.create_sheet("MAT Data")
                ws_mat['A1'] = "Market Analysis Data"
                ws_mat['A1'].font = Font(bold=True, size=14)
                
                row = 3
                for tool, data in st.session_state.mat_data.items():
                    ws_mat[f'A{row}'] = tool
                    ws_mat[f'A{row}'].font = Font(bold=True)
                    row += 1
                    
                    if isinstance(data, dict):
                        for key, value in data.items():
                            ws_mat[f'A{row}'] = str(key)
                            ws_mat[f'B{row}'] = str(value)
                            row += 1
                    row += 1
                
                # Save to BytesIO
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)
                
                st.download_button(
                    label="📥 Download Excel",
                    data=excel_buffer,
                    file_name=f"siama_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"Error creating Excel file: {e}")
    
    st.markdown("---")
    
    # Training Recommendations
    st.subheader("🎯 Training Program Recommendations")
    
    if st.button("Generate Recommendations"):
        st.markdown("""
        ### Based on Your SIAMA Analysis:
        
        #### **Key Findings:**
        """)
        
        # SIT findings
        if st.session_state.sit_data.get('roles'):
            st.markdown("**Stakeholder Landscape:**")
            for role, actors in st.session_state.sit_data['roles'].items():
                st.write(f"- {role}: {len(actors)} actors identified")
        
        # SAT findings
        if 'relationship_data' in st.session_state.sat_data:
            st.markdown("**Stakeholder Analysis:**")
            df = pd.DataFrame(st.session_state.sat_data['relationship_data'])
            high_power_high_interest = len(df[(df['power'] > 5) & (df['interest'] > 5)])
            st.write(f"- {high_power_high_interest} key stakeholders requiring close management")
        
        # MAT findings
        if st.session_state.mat_data:
            st.markdown("**Market Insights:**")
            st.write(f"- {len(st.session_state.mat_data)} market analysis tools completed")
        
        st.markdown("""
        #### **Recommended Training Focus Areas:**
        
        1. **Technical Skills Development**
           - Based on identified gaps in producer capabilities
           - Focus on quality improvement and efficiency
        
        2. **Business & Entrepreneurship**
           - Market understanding and customer engagement
           - Pricing strategies and financial management
        
        3. **Digital Literacy**
           - E-commerce platform usage
           - Social media marketing
           - Digital payment systems
        
        4. **Stakeholder Collaboration**
           - Building effective supplier relationships
           - Customer relationship management
           - Networking with marketers and buyers
        
        5. **Design & Innovation**
           - Contemporary market trends
           - Product diversification
           - Maintaining cultural authenticity while innovating
        """)

# Footer
st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray;'>
    <p>SIAMA Toolbox v1.0 | Developed for Craft Education Planning in India</p>
    <p>Based on research by Kumar et al., IIT Guwahati</p>
</div>
""", unsafe_allow_html=True)
